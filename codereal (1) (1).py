from flask import Flask, g, render_template_string, request, redirect, url_for, flash, send_file, jsonify, session
import sqlite3
import os
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
import csv
import io
import zipfile
from functools import wraps

# Importaciones opcionales con manejo de errores
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from twilio.rest import Client as TwilioClient
    TWILIO_AVAILABLE = True
except ImportError:
    TWILIO_AVAILABLE = False

# Configuración
DB_PATH = os.path.join(os.path.dirname(__file__), 'lavanderia.db')
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-12345')

# ---------------------- BASE DE DATOS ----------------------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    cur = db.cursor()
    
    # Tablas - ACTUALIZADA la tabla price_list
    tables = [
        '''CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT UNIQUE NOT NULL,
            address TEXT,
            created_at TEXT
        )''',
        '''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TEXT
        )''',
        '''CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            qty INTEGER NOT NULL DEFAULT 0,
            low_threshold INTEGER DEFAULT 5
        )''',
        '''CREATE TABLE IF NOT EXISTS price_list (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            garment_type TEXT UNIQUE NOT NULL,
            price REAL NOT NULL,
            category TEXT NOT NULL
        )''',
        '''CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE NOT NULL,
            client_id INTEGER,
            status TEXT NOT NULL,
            created_at TEXT,
            delivery_date TEXT,
            total REAL DEFAULT 0,
            notes TEXT,
            FOREIGN KEY(client_id) REFERENCES clients(id)
        )''',
        '''CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            garment_type TEXT,
            quantity INTEGER,
            unit_price REAL,
            subtotal REAL,
            FOREIGN KEY(order_id) REFERENCES orders(id)
        )''',
        '''CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            table_name TEXT,
            row_id INTEGER,
            username TEXT,
            created_at TEXT
        )'''
    ]
    
    for table in tables:
        cur.execute(table)
    
    db.commit()
    seed_defaults(db)

def seed_defaults(db):
    cur = db.cursor()
    
    # Usuario admin por defecto
    cur.execute("SELECT COUNT(*) as c FROM users")
    if cur.fetchone()['c'] == 0:
        cur.execute("INSERT INTO users (username,password_hash,role,created_at) VALUES (?,?,?,?)",
                    ('admin', generate_password_hash('admin123'), 'admin', datetime.utcnow().isoformat()))
    
    # PRENDAS TÍPICAS DE REPÚBLICA DOMINICANA POR CATEGORÍAS
    categories_and_prices = [
        # Ropa Casual
        ('camisa casual', 25.0, 'ropa_casual'),
        ('pantalón jean', 35.0, 'ropa_casual'),
        ('short', 20.0, 'ropa_casual'),
        ('blusa', 22.0, 'ropa_casual'),
        ('top', 15.0, 'ropa_casual'),
        
        # Ropa Formal
        ('camisa formal', 30.0, 'ropa_formal'),
        ('pantalón de vestir', 40.0, 'ropa_formal'),
        ('traje completo', 80.0, 'ropa_formal'),
        ('corbata', 10.0, 'ropa_formal'),
        ('chaleco', 25.0, 'ropa_formal'),
        
        # Ropa Deportiva
        ('uniforme deportivo', 35.0, 'deportiva'),
        ('short deportivo', 18.0, 'deportiva'),
        ('franela deportiva', 20.0, 'deportiva'),
        ('leggings', 22.0, 'deportiva'),
        ('sudadera', 45.0, 'deportiva'),
        
        # Ropa Interior
        ('calzoncillos', 8.0, 'interior'),
        ('panties', 6.0, 'interior'),
        ('sostén', 10.0, 'interior'),
        ('camiseta interior', 7.0, 'interior'),
        ('pijama', 30.0, 'interior'),
        
        # Ropa de Cama
        ('sábana individual', 50.0, 'cama'),
        ('sábana doble', 70.0, 'cama'),
        ('funda de almohada', 15.0, 'cama'),
        ('cobija', 120.0, 'cama'),
        ('edredón', 150.0, 'cama'),
        
        # Toallas y Manteles
        ('toalla de baño', 25.0, 'toallas'),
        ('toalla de mano', 12.0, 'toallas'),
        ('toalla de playa', 40.0, 'toallas'),
        ('mantel individual', 20.0, 'toallas'),
        ('mantel grande', 45.0, 'toallas'),
        
        # Uniformes
        ('uniforme escolar', 35.0, 'uniformes'),
        ('uniforme médico', 45.0, 'uniformes'),
        ('uniforme de trabajo', 40.0, 'uniformes'),
        ('delantal', 15.0, 'uniformes'),
        ('bata', 35.0, 'uniformes')
    ]
    
    for garment, price, category in categories_and_prices:
        cur.execute('INSERT OR IGNORE INTO price_list (garment_type, price, category) VALUES (?,?,?)', 
                   (garment, price, category))
    
    # Inventario por defecto
    items = [('detergente', 50, 5), ('suavizante', 40, 5), ('bolsas', 200, 10)]
    for name, qty, low in items:
        cur.execute('INSERT OR IGNORE INTO inventory (name,qty,low_threshold) VALUES (?,?,?)', (name, qty, low))
    
    db.commit()

# Inicializar BD al inicio
with app.app_context():
    init_db()

# ---------------------- UTILIDADES ----------------------
def log_action(action, table, row_id=None, username='system'):
    db = get_db()
    cur = db.cursor()
    cur.execute('INSERT INTO audit_logs (action,table_name,row_id,username,created_at) VALUES (?,?,?,?,?)',
                (action, table, row_id, username, datetime.utcnow().isoformat()))
    db.commit()

def generate_order_number():
    db = get_db()
    cur = db.cursor()
    # Buscar por el prefijo 'YYYY-MM-DD' que es lo que tiene created_at
    today_prefix = datetime.utcnow().date().isoformat()  # 'YYYY-MM-DD'
    cur.execute("SELECT COUNT(*) as c FROM orders WHERE created_at LIKE ?", (today_prefix + '%',))
    seq = cur.fetchone()['c'] + 1
    return f"{datetime.utcnow().strftime('%Y%m%d')}-{seq:04d}"


def get_garments_by_category():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT DISTINCT category FROM price_list ORDER BY category')
    categories = cur.fetchall()
    
    garments_by_category = {}
    for category in categories:
        cur.execute('SELECT garment_type, price FROM price_list WHERE category = ? ORDER BY garment_type', 
                   (category['category'],))
        garments_by_category[category['category']] = cur.fetchall()
    
    return garments_by_category

# ---------------------- AUTENTICACIÓN ----------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_role' not in session or session['user_role'] != 'admin':
            flash('Acceso denegado: permisos insuficientes', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        cur = db.cursor()
        cur.execute('SELECT * FROM users WHERE username=?', (username,))
        user = cur.fetchone()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['user_role'] = user['role']
            flash('Sesión iniciada correctamente', 'success')
            return redirect(url_for('index'))
        flash('Usuario o contraseña incorrectos', 'danger')
    return render_template_string(LOGIN_TEMPLATE)

@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada', 'info')
    return redirect(url_for('login'))

# ---------------------- RUTAS PRINCIPALES ----------------------
@app.route('/')
@login_required
def index():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT o.*, c.name as client_name FROM orders o LEFT JOIN clients c ON o.client_id=c.id ORDER BY o.created_at DESC LIMIT 20")
    orders = cur.fetchall()
    cur.execute('SELECT * FROM inventory WHERE qty <= low_threshold')
    low_items = cur.fetchall()
    return render_template_string(INDEX_TEMPLATE, orders=orders, low_items=low_items)

# ---------------------- CLIENTES ----------------------
@app.route('/clients')
@login_required
def clients():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT * FROM clients ORDER BY name')
    clients_list = cur.fetchall()
    return render_template_string(CLIENTS_TEMPLATE, clients=clients_list)

@app.route('/clients/new', methods=['GET','POST'])
@login_required
def new_client():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        address = request.form.get('address', '')
        db = get_db()
        cur = db.cursor()
        try:
            cur.execute('INSERT INTO clients (name,phone,address,created_at) VALUES (?,?,?,?)',
                        (name, phone, address, datetime.utcnow().isoformat()))
            db.commit()
            client_id = cur.lastrowid
            log_action('create_client', 'clients', client_id, session.get('username'))
            flash('Cliente creado exitosamente', 'success')
            return redirect(url_for('clients'))
        except sqlite3.IntegrityError:
            flash('Error: El teléfono ya existe', 'danger')
    return render_template_string(NEW_CLIENT_TEMPLATE)

@app.route('/clients/<int:client_id>')
@login_required
def client_detail(client_id):
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT * FROM clients WHERE id=?', (client_id,))
    client = cur.fetchone()
    cur.execute('SELECT * FROM orders WHERE client_id=? ORDER BY created_at DESC', (client_id,))
    orders = cur.fetchall()
    return render_template_string(CLIENT_DETAIL_TEMPLATE, client=client, orders=orders)

# ---------------------- INVENTARIO ----------------------
@app.route('/inventory')
@login_required
def inventory():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT * FROM inventory ORDER BY name')
    items = cur.fetchall()
    return render_template_string(INVENTORY_TEMPLATE, items=items)

@app.route('/inventory/edit/<int:item_id>', methods=['GET','POST'])
@admin_required
def edit_inventory(item_id):
    db = get_db()
    cur = db.cursor()
    if request.method == 'POST':
        qty = int(request.form['qty'])
        low = int(request.form['low_threshold'])
        cur.execute('UPDATE inventory SET qty=?, low_threshold=? WHERE id=?', (qty, low, item_id))
        db.commit()
        log_action('edit_inventory', 'inventory', item_id, session.get('username'))
        flash('Inventario actualizado', 'success')
        return redirect(url_for('inventory'))
    cur.execute('SELECT * FROM inventory WHERE id=?', (item_id,))
    item = cur.fetchone()
    return render_template_string(INVENTORY_EDIT_TEMPLATE, item=item)

# ---------------------- PRECIOS ----------------------
@app.route('/prices')
@login_required
def prices():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT * FROM price_list ORDER BY category, garment_type')
    prices_list = cur.fetchall()
    return render_template_string(PRICES_TEMPLATE, prices=prices_list)

@app.route('/prices/edit/<int:pid>', methods=['GET','POST'])
@admin_required
def edit_price(pid):
    db = get_db()
    cur = db.cursor()
    if request.method == 'POST':
        garment_type = request.form['garment_type']
        price = float(request.form['price'])
        category = request.form['category']
        cur.execute('UPDATE price_list SET garment_type=?, price=?, category=? WHERE id=?', 
                   (garment_type, price, category, pid))
        db.commit()
        log_action('edit_price', 'price_list', pid, session.get('username'))
        flash('Precio actualizado', 'success')
        return redirect(url_for('prices'))
    cur.execute('SELECT * FROM price_list WHERE id=?', (pid,))
    price_item = cur.fetchone()
    return render_template_string(PRICE_EDIT_TEMPLATE, p=price_item)

@app.route('/prices/new', methods=['GET','POST'])
@admin_required
def new_price():
    if request.method == 'POST':
        garment_type = request.form['garment_type']
        price = float(request.form['price'])
        category = request.form['category']
        db = get_db()
        cur = db.cursor()
        try:
            cur.execute('INSERT INTO price_list (garment_type, price, category) VALUES (?,?,?)',
                        (garment_type, price, category))
            db.commit()
            flash('Prenda agregada exitosamente', 'success')
            return redirect(url_for('prices'))
        except sqlite3.IntegrityError:
            flash('Error: Ya existe una prenda con ese nombre', 'danger')
    return render_template_string(NEW_PRICE_TEMPLATE)

# ---------------------- ÓRDENES ----------------------
@app.route('/orders/new', methods=['GET','POST'])
@login_required
def new_order():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT * FROM clients ORDER BY name')
    clients_list = cur.fetchall()
    
    # Obtener prendas por categoría
    garments_by_category = get_garments_by_category()
    
    if request.method == 'POST':
        client_id = int(request.form['client_id']) if request.form.get('client_id') else None
        delivery_date = request.form['delivery_date']
        notes = request.form.get('notes', '')
        
        # Generar número de orden de manera segura
        order_number = generate_order_number()
        created_at = datetime.utcnow().isoformat()
        
        # Verificar que el número de orden no existe (double-check)
        cur.execute('SELECT id FROM orders WHERE order_number = ?', (order_number,))
        if cur.fetchone():
            # Si por alguna razón ya existe, generar uno nuevo
            order_number = generate_order_number()
        
        try:
            cur.execute('INSERT INTO orders (order_number,client_id,status,created_at,delivery_date,notes) VALUES (?,?,?,?,?,?)',
                        (order_number, client_id, 'pendiente', created_at, delivery_date, notes))
            order_id = cur.lastrowid
            total = 0.0
            
            # Procesar items
            items_added = False
            for key in request.form:
                if key.startswith('qty_'):
                    garment = key.split('qty_')[1]
                    qty = int(request.form[key] or 0)
                    if qty <= 0:
                        continue
                    
                    cur.execute('SELECT price FROM price_list WHERE garment_type=?', (garment,))
                    row = cur.fetchone()
                    unit_price = row['price'] if row else 0.0
                    subtotal = unit_price * qty
                    total += subtotal
                    
                    cur.execute('INSERT INTO order_items (order_id,garment_type,quantity,unit_price,subtotal) VALUES (?,?,?,?,?)',
                                (order_id, garment, qty, unit_price, subtotal))
                    items_added = True
            
            # Si no se agregaron items, eliminar la orden vacía
            if not items_added:
                cur.execute('DELETE FROM orders WHERE id=?', (order_id,))
                db.commit()
                flash('Error: La orden debe tener al menos una prenda', 'danger')
                return redirect(url_for('new_order'))
            
            cur.execute('UPDATE orders SET total=? WHERE id=?', (total, order_id))
            db.commit()
            log_action('create_order', 'orders', order_id, session.get('username'))
            flash(f'Orden {order_number} creada exitosamente', 'success')
            return redirect(url_for('index'))
        
        except sqlite3.IntegrityError as e:
            db.rollback()
            if 'UNIQUE constraint failed: orders.order_number' in str(e):
                # Regenerar número y reintentar (solo una vez)
                try:
                    order_number = generate_order_number()
                    cur.execute('INSERT INTO orders (order_number,client_id,status,created_at,delivery_date,notes) VALUES (?,?,?,?,?,?)',
                                (order_number, client_id, 'pendiente', created_at, delivery_date, notes))
                    # ... resto del código de inserción
                    db.commit()
                    flash(f'Orden {order_number} creada exitosamente', 'success')
                    return redirect(url_for('index'))
                except Exception as e2:
                    db.rollback()
                    flash(f'Error crítico al crear orden: {str(e2)}', 'danger')
            else:
                flash(f'Error al crear orden: {str(e)}', 'danger')
    
    # Pasar la fecha de hoy al template
    today = date.today().isoformat()
    
    return render_template_string(NEW_ORDER_TEMPLATE, 
                                 clients=clients_list, 
                                 garments_by_category=garments_by_category,
                                 today=today)

@app.route('/orders/<int:order_id>')
@login_required
def order_detail(order_id):
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT o.*, c.name as client_name, c.phone FROM orders o LEFT JOIN clients c ON o.client_id=c.id WHERE o.id=?', (order_id,))
    order = cur.fetchone()
    cur.execute('SELECT * FROM order_items WHERE order_id=?', (order_id,))
    items = cur.fetchall()
    return render_template_string(ORDER_DETAIL_TEMPLATE, order=order, items=items)

@app.route('/orders/<int:order_id>/status', methods=['POST'])
@login_required
def change_status(order_id):
    status = request.form['status']
    db = get_db()
    cur = db.cursor()
    cur.execute('UPDATE orders SET status=? WHERE id=?', (status, order_id))
    db.commit()
    log_action('change_status', 'orders', order_id, session.get('username'))
    
    if status == 'listo':
        try:
            send_notification(order_id)
            flash('Notificación enviada al cliente', 'info')
        except Exception as e:
            flash(f'Error al enviar notificación: {str(e)}', 'warning')
    
    flash('Estado actualizado', 'success')
    return redirect(url_for('order_detail', order_id=order_id))

# ---------------------- REPORTES Y EXPORTACIÓN ----------------------
@app.route('/reports')
@login_required
def reports():
    db = get_db()
    cur = db.cursor()
    today = date.today().isoformat()
    cur.execute("SELECT SUM(total) as total FROM orders WHERE created_at LIKE ?", (today + '%',))
    sales_today = cur.fetchone()['total'] or 0
    cur.execute("SELECT garment_type, SUM(quantity) as q FROM order_items GROUP BY garment_type ORDER BY q DESC LIMIT 10")
    popular = cur.fetchall()
    return render_template_string(REPORTS_TEMPLATE, sales_today=sales_today, popular=popular)

@app.route('/export/orders.csv')
@login_required
def export_orders_csv():
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT o.*, c.name as client_name, c.phone FROM orders o LEFT JOIN clients c ON o.client_id=c.id')
    rows = cur.fetchall()
    
    output = io.BytesIO()
    writer = csv.writer(io.TextIOWrapper(output, encoding='utf-8'))
    writer.writerow(['order_number','client_name','phone','status','created_at','delivery_date','total','notes'])
    
    for r in rows:
        writer.writerow([r['order_number'], r['client_name'], r['phone'], r['status'], r['created_at'], r['delivery_date'], r['total'], r['notes']])
    
    output.seek(0)
    return send_file(output, mimetype='text/csv', as_attachment=True, download_name='orders.csv')

@app.route('/export/orders.xlsx')
@login_required
def export_orders_xlsx():
    if not OPENPYXL_AVAILABLE:
        flash('openpyxl no está instalado. Instala con: pip install openpyxl', 'danger')
        return redirect(url_for('reports'))
    
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT o.*, c.name as client_name, c.phone FROM orders o LEFT JOIN clients c ON o.client_id=c.id')
    rows = cur.fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes'
    headers = ['Número Orden','Cliente','Teléfono','Estado','Fecha Creación','Fecha Entrega','Total','Notas']
    ws.append(headers)
    
    for r in rows:
        ws.append([r['order_number'], r['client_name'], r['phone'], r['status'], r['created_at'], r['delivery_date'], r['total'], r['notes']])
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    as_attachment=True, download_name='ordenes.xlsx')

@app.route('/export/backup_all.zip')
@admin_required
def export_backup_zip():
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode='w') as z:
        db = get_db()
        cur = db.cursor()
        
        # Clientes
        cur.execute('SELECT * FROM clients')
        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(['id','name','phone','address','created_at'])
        for r in cur.fetchall():
            w.writerow([r['id'], r['name'], r['phone'], r['address'], r['created_at']])
        z.writestr('clients.csv', si.getvalue())
        
        # Órdenes
        cur.execute('SELECT * FROM orders')
        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(['id','order_number','client_id','status','created_at','delivery_date','total','notes'])
        for r in cur.fetchall():
            w.writerow([r['id'], r['order_number'], r['client_id'], r['status'], r['created_at'], r['delivery_date'], r['total'], r['notes']])
        z.writestr('orders.csv', si.getvalue())
        
        # Inventario
        cur.execute('SELECT * FROM inventory')
        si = io.StringIO()
        w = csv.writer(si)
        w.writerow(['id','name','qty','low_threshold'])
        for r in cur.fetchall():
            w.writerow([r['id'], r['name'], r['qty'], r['low_threshold']])
        z.writestr('inventory.csv', si.getvalue())
    
    mem.seek(0)
    return send_file(mem, mimetype='application/zip', as_attachment=True, download_name='backup.zip')

# ---------------------- NOTIFICACIONES TWILIO ----------------------
def get_twilio_client():
    sid = os.environ.get('TWILIO_ACCOUNT_SID')
    token = os.environ.get('TWILIO_AUTH_TOKEN')
    if not sid or not token or not TWILIO_AVAILABLE:
        return None
    return TwilioClient(sid, token)

def send_notification(order_id, channel='sms'):
    """Enviar notificación SMS/WhatsApp al cliente"""
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT o.order_number, c.phone, c.name FROM orders o LEFT JOIN clients c ON o.client_id=c.id WHERE o.id=?', (order_id,))
    result = cur.fetchone()
    
    if not result:
        raise ValueError('Orden no encontrada')
    
    phone = result['phone']
    client_name = result['name'] or 'cliente'
    order_number = result['order_number']
    
    tw_client = get_twilio_client()
    if not tw_client:
        raise RuntimeError('Twilio no configurado')
    
    from_number = os.environ.get('TWILIO_FROM_NUMBER')
    if channel == 'whatsapp':
        from_number = os.environ.get('TWILIO_WHATSAPP_FROM', from_number)
        to_number = f'whatsapp:{phone}' if not phone.startswith('whatsapp:') else phone
    else:
        to_number = phone
    
    body = f'Hola {client_name}, su orden {order_number} ya está lista. ¡Gracias por preferirnos!'
    
    try:
        msg = tw_client.messages.create(body=body, from_=from_number, to=to_number)
        log_action('send_notification', 'orders', order_id, session.get('username','system'))
        return msg.sid
    except Exception as e:
        raise RuntimeError(f'Error enviando mensaje: {str(e)}')

# ---------------------- GENERACIÓN DE PDF ----------------------
def generate_receipt_pdf(order_id):
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError('reportlab no disponible. Instala con: pip install reportlab')
    
    db = get_db()
    cur = db.cursor()
    cur.execute('SELECT o.*, c.name as client_name, c.phone, c.address FROM orders o LEFT JOIN clients c ON o.client_id=c.id WHERE o.id=?', (order_id,))
    order = cur.fetchone()
    
    if not order:
        raise ValueError('Orden no encontrada')
    
    cur.execute('SELECT * FROM order_items WHERE order_id=?', (order_id,))
    items = cur.fetchall()
    
    bio = io.BytesIO()
    c = pdf_canvas.Canvas(bio, pagesize=letter)
    width, height = letter
    
    # Encabezado
    company = os.environ.get('COMPANY_NAME', 'Lavandería Effiwash')
    logo_path = os.environ.get('COMPANY_LOGO_PATH')
    y = height - 50
    
    # Logo (opcional)
    if logo_path and os.path.exists(logo_path):
        try:
            c.drawImage(logo_path, 40, y-60, width=120, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    
    c.setFont('Helvetica-Bold', 16)
    c.drawString(180, y, company)
    c.setFont('Helvetica', 10)
    c.drawString(180, y-15, f'Orden: {order["order_number"]}')
    c.drawString(180, y-30, f'Fecha: {order["created_at"][:19]}')
    
    # Información del cliente
    c.drawString(40, y-80, f'Cliente: {order["client_name"] or "-"}')
    c.drawString(40, y-95, f'Teléfono: {order["phone"] or "-"}')
    c.drawString(40, y-110, f'Entrega: {order["delivery_date"] or "-"}')
    
    # Tabla de items
    start_y = y-140
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, start_y, 'Prenda')
    c.drawString(200, start_y, 'Cantidad')
    c.drawString(260, start_y, 'P. Unit.')
    c.drawString(320, start_y, 'Subtotal')
    
    c.setFont('Helvetica', 10)
    yy = start_y - 20
    
    for item in items:
        c.drawString(40, yy, item['garment_type'])
        c.drawString(200, yy, str(item['quantity']))
        c.drawString(260, yy, f"${item['unit_price']:.2f}")
        c.drawString(320, yy, f"${item['subtotal']:.2f}")
        yy -= 15
        
        if yy < 100:  # Nueva página si es necesario
            c.showPage()
            yy = height - 50
    
    # Total
    c.setFont('Helvetica-Bold', 12)
    c.drawString(260, yy-20, 'TOTAL:')
    c.drawString(320, yy-20, f"${order['total']:.2f}")
    
    c.save()
    bio.seek(0)
    return bio

@app.route('/orders/<int:order_id>/receipt')
@login_required
def order_receipt(order_id):
    try:
        pdf_io = generate_receipt_pdf(order_id)
        return send_file(pdf_io, mimetype='application/pdf', 
                        as_attachment=True, download_name=f'recibo_{order_id}.pdf')
    except Exception as e:
        flash(f'Error generando PDF: {str(e)}', 'danger')
        return redirect(url_for('order_detail', order_id=order_id))

# ---------------------- TEMPLATES COMPLETOS ----------------------
BASE_HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sistema de Lavandería Effiwash</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .navbar-brand { font-weight: bold; }
        .table-hover tbody tr:hover { background-color: #f5f5f5; }
        .alert { margin-top: 20px; }
        .sticky-info { display: flex; top: 20px; z-index: 100; }
    </style>
</head>
<body class="bg-light">
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark">
        <div class="container">
            <a class="navbar-brand" href="/"> Lavandería Effiwash Express</a>
            <div class="navbar-nav ms-auto">
                <span class="navbar-text me-3">Usuario: {{ session.username }}</span>
                <a class="btn btn-outline-light btn-sm" href="/logout">Cerrar Sesión</a>
            </div>
        </div>
    </nav>

    <div class="container mt-4">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show">
                        {{ message }}
                        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        {% block content %}{% endblock %}
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Iniciar Sesión - Lavandería Effiwash</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light d-flex align-items-center justify-content-center" style="min-height: 100vh;">
    <div class="card shadow" style="width: 100%; max-width: 400px;">
        <div class="card-body p-4">
            <h3 class="card-title text-center mb-4"> Iniciar Sesión</h3>
            
            <form method="post">
                <div class="mb-3">
                    <label class="form-label">Usuario:</label>
                    <input type="text" class="form-control" name="username" required>
                </div>
                
                <div class="mb-3">
                    <label class="form-label">Contraseña:</label>
                    <input type="password" class="form-control" name="password" required>
                </div>
                
                <button type="submit" class="btn btn-primary w-100">Entrar al Sistema</button>
            </form>
            
            <div class="mt-3 text-center">
                <small class="text-muted">Usuario demo: admin / admin123</small>
            </div>
        </div>
    </div>
</body>
</html>
"""

INDEX_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row">
    <div class="col-md-8">
        <h2> Órdenes Recientes</h2>
        <div class="table-responsive">
            <table class="table table-striped table-hover">
                <thead class="table-dark">
                    <tr>
                        <th># Orden</th>
                        <th>Cliente</th>
                        <th>Estado</th>
                        <th>Total</th>
                        <th>Acciones</th>
                    </tr>
                </thead>
                <tbody>
                    {% for order in orders %}
                    <tr>
                        <td>{{ order.order_number }}</td>
                        <td>{{ order.client_name or 'No especificado' }}</td>
                        <td>
                            <span class="badge bg-{{ 'success' if order.status == 'listo' else 'warning' if order.status == 'pendiente' else 'secondary' }}">
                                {{ order.status }}
                            </span>
                        </td>
                        <td>${{ "%.2f"|format(order.total or 0) }}</td>
                        <td>
                            <a href="/orders/{{ order.id }}" class="btn btn-sm btn-info">Ver</a>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </div>
    
    <div class="col-md-4">
        <div class="card">
            <div class="card-header bg-warning">
                <h5 class="mb-0"> Alertas de Inventario</h5>
            </div>
            <div class="card-body">
                {% if low_items %}
                    <ul class="list-group">
                        {% for item in low_items %}
                        <li class="list-group-item d-flex justify-content-between align-items-center">
                            {{ item.name }}
                            <span class="badge bg-danger rounded-pill">{{ item.qty }}</span>
                        </li>
                        {% endfor %}
                    </ul>
                {% else %}
                    <p class="text-muted">No hay alertas de inventario</p>
                {% endif %}
            </div>
        </div>
        
        <div class="card mt-3">
            <div class="card-header bg-primary text-white">
                <h5 class="mb-0"> Acciones Rápidas</h5>
            </div>
            <div class="card-body">
                <a href="/orders/new" class="btn btn-success w-100 mb-2"> Nueva Orden</a>
                <a href="/clients" class="btn btn-outline-primary w-100 mb-2"> Gestionar Clientes</a>
                <a href="/reports" class="btn btn-outline-info w-100"> Ver Reportes</a>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

CLIENTS_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Gestión de Clientes</h2>
    <a href="/clients/new" class="btn btn-success"> Nuevo Cliente</a>
</div>

<div class="table-responsive">
    <table class="table table-striped table-hover">
        <thead class="table-dark">
            <tr>
                <th>Nombre</th>
                <th>Teléfono</th>
                <th>Dirección</th>
                <th>Acciones</th>
            </tr>
        </thead>
        <tbody>
            {% for client in clients %}
            <tr>
                <td>{{ client.name }}</td>
                <td>{{ client.phone }}</td>
                <td>{{ client.address or '-' }}</td>
                <td>
                    <a href="/clients/{{ client.id }}" class="btn btn-sm btn-info"> Ver Órdenes</a>
                </td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% endblock %}
''')

NEW_CLIENT_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-6">
        <h2> Nuevo Cliente</h2>
        <div class="card">
            <div class="card-body">
                <form method="post">
                    <div class="mb-3">
                        <label class="form-label">Nombre completo:</label>
                        <input type="text" class="form-control" name="name" required>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Teléfono:</label>
                        <input type="text" class="form-control" name="phone" required>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Dirección (opcional):</label>
                        <textarea class="form-control" name="address" rows="3"></textarea>
                    </div>
                    <div class="d-grid gap-2">
                        <button type="submit" class="btn btn-primary">Guardar Cliente</button>
                        <a href="/clients" class="btn btn-secondary">Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

CLIENT_DETAIL_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Detalles del Cliente</h2>
    <a href="/clients" class="btn btn-secondary">← Volver a Clientes</a>
</div>

<div class="row">
    <div class="col-md-6">
        <div class="card">
            <div class="card-header">
                <h5 class="mb-0">Información del Cliente</h5>
            </div>
            <div class="card-body">
                <p><strong>Nombre:</strong> {{ client.name }}</p>
                <p><strong>Teléfono:</strong> {{ client.phone }}</p>
                <p><strong>Dirección:</strong> {{ client.address or 'No especificada' }}</p>
                <p><strong>Fecha de registro:</strong> {{ client.created_at[:10] }}</p>
            </div>
        </div>
    </div>
    
    <div class="col-md-6">
        <div class="card">
            <div class="card-header">
                <h5 class="mb-0">Órdenes del Cliente</h5>
            </div>
            <div class="card-body">
                {% if orders %}
                    <div class="list-group">
                        {% for order in orders %}
                        <a href="/orders/{{ order.id }}" class="list-group-item list-group-item-action">
                            <div class="d-flex w-100 justify-content-between">
                                <h6 class="mb-1">Orden #{{ order.order_number }}</h6>
                                <span class="badge bg-{{ 'success' if order.status == 'listo' else 'warning' if order.status == 'pendiente' else 'secondary' }}">
                                    {{ order.status }}
                                </span>
                            </div>
                            <p class="mb-1">Total: ${{ "%.2f"|format(order.total or 0) }}</p>
                            <small>Entrega: {{ order.delivery_date or 'No especificada' }}</small>
                        </a>
                        {% endfor %}
                    </div>
                {% else %}
                    <p class="text-muted">Este cliente no tiene órdenes registradas.</p>
                {% endif %}
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

INVENTORY_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Gestión de Inventario</h2>
    {% if session.user_role == 'admin' %}
    <a href="/inventory/edit/0" class="btn btn-success"> Nuevo Producto</a>
    {% endif %}
</div>

<div class="table-responsive">
    <table class="table table-striped table-hover">
        <thead class="table-dark">
            <tr>
                <th>Producto</th>
                <th>Cantidad</th>
                <th>Umbral Mínimo</th>
                <th>Estado</th>
                {% if session.user_role == 'admin' %}
                <th>Acciones</th>
                {% endif %}
            </tr>
        </thead>
        <tbody>
            {% for item in items %}
            <tr>
                <td>{{ item.name }}</td>
                <td>{{ item.qty }}</td>
                <td>{{ item.low_threshold }}</td>
                <td>
                    {% if item.qty <= item.low_threshold %}
                    <span class="badge bg-danger">Bajo Stock</span>
                    {% else %}
                    <span class="badge bg-success">Disponible</span>
                    {% endif %}
                </td>
                {% if session.user_role == 'admin' %}
                <td>
                    <a href="/inventory/edit/{{ item.id }}" class="btn btn-sm btn-warning">Editar</a>
                </td>
                {% endif %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% endblock %}
''')

INVENTORY_EDIT_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-6">
        <h2>{% if item %} Editar Producto{% else %} Nuevo Producto{% endif %}</h2>
        <div class="card">
            <div class="card-body">
                <form method="post">
                    <div class="mb-3">
                        <label class="form-label">Nombre del producto:</label>
                        <input type="text" class="form-control" name="name" value="{{ item.name if item else '' }}" required>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Cantidad en stock:</label>
                        <input type="number" class="form-control" name="qty" value="{{ item.qty if item else 0 }}" required min="0">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Umbral mínimo (alerta):</label>
                        <input type="number" class="form-control" name="low_threshold" value="{{ item.low_threshold if item else 5 }}" required min="1">
                    </div>
                    <div class="d-grid gap-2">
                        <button type="submit" class="btn btn-primary">{% if item %}Actualizar{% else %}Crear{% endif %} Producto</button>
                        <a href="/inventory" class="btn btn-secondary">Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

PRICES_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Lista de Precios</h2>
    {% if session.user_role == 'admin' %}
    <a href="/prices/new" class="btn btn-success"> Nueva Prenda</a>
    {% endif %}
</div>

<div class="table-responsive">
    <table class="table table-striped table-hover">
        <thead class="table-dark">
            <tr>
                <th>Categoría</th>
                <th>Tipo de Prenda</th>
                <th>Precio Unitario</th>
                {% if session.user_role == 'admin' %}
                <th>Acciones</th>
                {% endif %}
            </tr>
        </thead>
        <tbody>
            {% for price in prices %}
            <tr>
                <td><span class="badge bg-info">{{ price.category.replace('_', ' ').title() }}</span></td>
                <td>{{ price.garment_type.title() }}</td>
                <td>${{ "%.2f"|format(price.price) }}</td>
                {% if session.user_role == 'admin' %}
                <td>
                    <a href="/prices/edit/{{ price.id }}" class="btn btn-sm btn-warning">Editar</a>
                </td>
                {% endif %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% endblock %}
''')

PRICE_EDIT_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-6">
        <h2>{% if p %} Editar Prenda{% else %} Nueva Prenda{% endif %}</h2>
        <div class="card">
            <div class="card-body">
                <form method="post">
                    <div class="mb-3">
                        <label class="form-label">Tipo de prenda:</label>
                        <input type="text" class="form-control" name="garment_type" value="{{ p.garment_type if p else '' }}" required>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Precio unitario ($):</label>
                        <input type="number" step="0.01" class="form-control" name="price" value="{{ "%.2f"|format(p.price) if p else '0.00' }}" required min="0">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Categoría:</label>
                        <select class="form-select" name="category" required>
                            <option value="">Seleccionar categoría</option>
                            <option value="ropa_casual" {{ 'selected' if p and p.category == 'ropa_casual' }}>Ropa Casual</option>
                            <option value="ropa_formal" {{ 'selected' if p and p.category == 'ropa_formal' }}>Ropa Formal</option>
                            <option value="deportiva" {{ 'selected' if p and p.category == 'deportiva' }}>Ropa Deportiva</option>
                            <option value="interior" {{ 'selected' if p and p.category == 'interior' }}>Ropa Interior</option>
                            <option value="cama" {{ 'selected' if p and p.category == 'cama' }}>Ropa de Cama</option>
                            <option value="toallas" {{ 'selected' if p and p.category == 'toallas' }}>Toallas y Manteles</option>
                            <option value="uniformes" {{ 'selected' if p and p.category == 'uniformes' }}>Uniformes</option>
                        </select>
                    </div>
                    <div class="d-grid gap-2">
                        <button type="submit" class="btn btn-primary">{% if p %}Actualizar{% else %}Crear{% endif %} Prenda</button>
                        <a href="/prices" class="btn btn-secondary">Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

NEW_PRICE_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-6">
        <h2> Nueva Prenda</h2>
        <div class="card">
            <div class="card-body">
                <form method="post">
                    <div class="mb-3">
                        <label class="form-label">Tipo de prenda:</label>
                        <input type="text" class="form-control" name="garment_type" required>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Precio unitario ($):</label>
                        <input type="number" step="0.01" class="form-control" name="price" required min="0">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Categoría:</label>
                        <select class="form-select" name="category" required>
                            <option value="">Seleccionar categoría</option>
                            <option value="ropa_casual">Ropa Casual</option>
                            <option value="ropa_formal">Ropa Formal</option>
                            <option value="deportiva">Ropa Deportiva</option>
                            <option value="interior">Ropa Interior</option>
                            <option value="cama">Ropa de Cama</option>
                            <option value="toallas">Toallas y Manteles</option>
                            <option value="uniformes">Uniformes</option>
                        </select>
                    </div>
                    <div class="d-grid gap-2">
                        <button type="submit" class="btn btn-primary">Crear Prenda</button>
                        <a href="/prices" class="btn btn-secondary">Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

# ---------------------- NUEVO TEMPLATE DE ORDEN CORREGIDO ----------------------
NEW_ORDER_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="row">
    <div class="col-md-8">
        <h2> 🧺 Nueva Orden de Lavado</h2>
        <div class="card">
            <div class="card-body">
                <form method="post" id="orderForm">
                    <div class="row mb-3">
                        <div class="col-md-6">
                            <label class="form-label">👤 Cliente (opcional):</label>
                            <select class="form-select" name="client_id" id="clientSelect">
                                <option value="">-- Seleccionar cliente --</option>
                                {% for client in clients %}
                                <option value="{{ client.id }}">{{ client.name }} - {{ client.phone }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="col-md-6">
                            <label class="form-label">📅 Fecha de entrega:</label>
                            <input type="date" class="form-control" name="delivery_date" required 
                                   min="{{ today }}" value="{{ today }}">
                        </div>
                    </div>
                    
                    <div class="mb-3">
                        <label class="form-label">📝 Notas (opcional):</label>
                        <textarea class="form-control" name="notes" rows="2" placeholder="Observaciones especiales..."></textarea>
                    </div>
                    
                    <!-- SECCIÓN DE SELECCIÓN DE PRENDAS -->
                    <div class="mb-4">
                        <h5>👕 Seleccionar Prendas a Lavar:</h5>
                        <div class="alert alert-info">
                            <small>💡 Haz clic en cada categoría para ver las prendas disponibles</small>
                        </div>
                        
                        <!-- NAVEGACIÓN POR CATEGORÍAS -->
                        <div class="row mb-3" id="categoryNav">
                            {% for category, garments in garments_by_category.items() %}
                            <div class="col-md-3 mb-2">
                                <button type="button" class="btn btn-outline-primary w-100 category-btn" 
                                        data-category="{{ category }}">
                                    {{ category.replace('_', ' ').title() }}
                                    <br><small>({{ garments|length }} prendas)</small>
                                </button>
                            </div>
                            {% endfor %}
                        </div>
                        
                        <!-- CONTENEDOR DE PRENDAS POR CATEGORÍA -->
                        <div id="garmentsContainer">
                            {% for category, garments in garments_by_category.items() %}
                            <div class="category-section" id="{{ category }}" style="display: none;">
                                <div class="card">
                                    <div class="card-header bg-primary text-white d-flex justify-content-between align-items-center">
                                        <h6 class="mb-0">{{ category.replace('_', ' ').title() }}</h6>
                                        <button type="button" class="btn btn-sm btn-light back-btn">← Volver</button>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            {% for garment in garments %}
                                            <div class="col-md-6 mb-3">
                                                <div class="card h-100">
                                                    <div class="card-body">
                                                        <h6 class="card-title">{{ garment.garment_type.title() }}</h6>
                                                        <p class="card-text text-success">Precio: ${{ "%.2f"|format(garment.price) }}</p>
                                                        <div class="input-group">
                                                            <button type="button" class="btn btn-outline-secondary btn-sm decrement" 
                                                                    data-garment="{{ garment.garment_type }}">-</button>
                                                            <input type="number" class="form-control form-control-sm text-center quantity-input" 
                                                                   name="qty_{{ garment.garment_type }}" value="0" min="0" 
                                                                   data-price="{{ garment.price }}" 
                                                                   data-garment="{{ garment.garment_type }}"
                                                                   readonly>
                                                            <button type="button" class="btn btn-outline-secondary btn-sm increment" 
                                                                    data-garment="{{ garment.garment_type }}">+</button>
                                                        </div>
                                                        <small class="text-muted d-block mt-1 subtotal" 
                                                               data-garment="{{ garment.garment_type }}">Subtotal: $0.00</small>
                                                    </div>
                                                </div>
                                            </div>
                                            {% endfor %}
                                        </div>
                                    </div>
                                </div>
                            </div>
                            {% endfor %}
                        </div>
                    </div>
                    
                    <!-- RESUMEN DE LA ORDEN -->
                    <div class="card mt-4">
                        <div class="card-header bg-success text-white">
                            <h5 class="mb-0">📊 Resumen de la Orden</h5>
                        </div>
                        <div class="card-body">
                            <div id="orderSummary">
                                <p class="text-muted">No hay prendas seleccionadas</p>
                            </div>
                            <div class="d-flex justify-content-between align-items-center mt-3">
                                <h4>Total: <span id="totalAmount">$0.00</span></h4>
                            </div>
                        </div>
                    </div>
                    
                    <div class="d-grid gap-2 mt-4">
                        <button type="submit" class="btn btn-success btn-lg">✅ Crear Orden</button>
                        <a href="/" class="btn btn-secondary">❌ Cancelar</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
    
    <div class="col-md-4">
        <div class="card sticky-info">
            <div class="card-header bg-info text-white">
                <h5 class="mb-0"> 💡 Información</h5>
            </div>
            <div class="card-body">
                <p><strong>Instrucciones:</strong></p>
                <ol class="small">
                    <li>Selecciona una categoría de prendas</li>
                    <li>Elige las cantidades con los botones +/‑</li>
                    <li>Usa "Volver" para cambiar de categoría</li>
                    <li>Revisa el resumen antes de crear la orden</li>
                </ol>
                <p class="mt-3"><strong>Características:</strong></p>
                <ul class="small">
                    <li>Prendas típicas</li>
                    <li>Precios en Dólares(US$)</li>
                    <li>Selección intuitiva por categorías</li>
                    <li>Cálculo automático del total</li>
                </ul>
            </div>
        </div>
        
        <!-- PRENDAS SELECCIONADAS -->
        <div class="card mt-3">
            <div class="card-header">
                <h6 class="mb-0">🛒 Prendas Seleccionadas</h6>
            </div>
            <div class="card-body">
                <div id="selectedItems" class="small">
                    <p class="text-muted">No hay prendas seleccionadas</p>
                </div>
            </div>
        </div>
    </div>
</div>

<style>
.category-section { animation: fadeIn 0.3s; }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
.quantity-input { max-width: 70px; }
.card:hover { transform: translateY(-2px); transition: transform 0.2s; }
</style>

<script>
// Variables globales
let selectedItems = {};

// Mostrar categoría seleccionada
document.querySelectorAll('.category-btn').forEach(btn => {
    btn.addEventListener('click', function() {
        const category = this.dataset.category;
        document.querySelectorAll('.category-section').forEach(section => {
            section.style.display = 'none';
        });
        document.getElementById(category).style.display = 'block';
        document.getElementById('categoryNav').style.display = 'none';
    });
});

// Botón volver
document.querySelectorAll('.back-btn').forEach(btn => {
    btn.addEventListener('click', function() {
        document.querySelectorAll('.category-section').forEach(section => {
            section.style.display = 'none';
        });
        document.getElementById('categoryNav').style.display = 'flex';
    });
});

// Incrementar cantidad
document.querySelectorAll('.increment').forEach(btn => {
    btn.addEventListener('click', function() {
        const garment = this.dataset.garment;
        const input = document.querySelector(`input[name="qty_${garment}"]`);
        input.value = parseInt(input.value) + 1;
        updateQuantity(garment, parseInt(input.value));
    });
});

// Decrementar cantidad
document.querySelectorAll('.decrement').forEach(btn => {
    btn.addEventListener('click', function() {
        const garment = this.dataset.garment;
        const input = document.querySelector(`input[name="qty_${garment}"]`);
        if (parseInt(input.value) > 0) {
            input.value = parseInt(input.value) - 1;
            updateQuantity(garment, parseInt(input.value));
        }
    });
});

// Actualizar cantidad y cálculos
function updateQuantity(garment, quantity) {
    const price = parseFloat(document.querySelector(`input[name="qty_${garment}"]`).dataset.price);
    const subtotal = price * quantity;
    
    // Actualizar subtotal en la tarjeta
    const subtotalElement = document.querySelector(`.subtotal[data-garment="${garment}"]`);
    if (subtotalElement) {
        subtotalElement.textContent = `Subtotal: $${subtotal.toFixed(2)}`;
    }
    
    // Actualizar items seleccionados
    if (quantity > 0) {
        selectedItems[garment] = { quantity, price, subtotal };
    } else {
        delete selectedItems[garment];
    }
    
    updateOrderSummary();
}

// Actualizar resumen de la orden
function updateOrderSummary() {
    const summaryDiv = document.getElementById('orderSummary');
    const selectedDiv = document.getElementById('selectedItems');
    const totalSpan = document.getElementById('totalAmount');
    
    let total = 0;
    
    if (Object.keys(selectedItems).length === 0) {
        summaryDiv.innerHTML = '<p class="text-muted">No hay prendas seleccionadas</p>';
        selectedDiv.innerHTML = '<p class="text-muted">No hay prendas seleccionadas</p>';
    } else {
        let summaryHTML = '<div class="table-responsive"><table class="table table-sm"><thead><tr><th>Prenda</th><th>Cant</th><th>Subtotal</th></tr></thead><tbody>';
        let selectedHTML = '<ul class="list-unstyled">';
        
        for (const [garment, data] of Object.entries(selectedItems)) {
            summaryHTML += `<tr><td>${garment.replace(/_/g, ' ').title()}</td><td>${data.quantity}</td><td>$${data.subtotal.toFixed(2)}</td></tr>`;
            selectedHTML += `<li class="mb-1">${garment.replace(/_/g, ' ').title()} × ${data.quantity} = $${data.subtotal.toFixed(2)}</li>`;
            total += data.subtotal;
        }
        
        summaryHTML += '</tbody></table></div>';
        selectedHTML += '</ul>';
        
        summaryDiv.innerHTML = summaryHTML;
        selectedDiv.innerHTML = selectedHTML;
    }
    
    totalSpan.textContent = `$${total.toFixed(2)}`;
}

// Función para capitalizar texto (title case)
String.prototype.title = function() {
    return this.replace(/\w\S*/g, function(txt) {
        return txt.charAt(0).toUpperCase() + txt.substr(1).toLowerCase();
    });
};

// Fecha mínima para entrega (hoy)
document.addEventListener('DOMContentLoaded', function() {
    const today = new Date().toISOString().split('T')[0];
    const deliveryInput = document.querySelector('input[name="delivery_date"]');
    if (deliveryInput) {
        deliveryInput.min = today;
        if (!deliveryInput.value) {
            deliveryInput.value = today;
        }
    }
    
    // Inicializar todos los inputs para que tengan evento de cambio
    document.querySelectorAll('.quantity-input').forEach(input => {
        input.addEventListener('change', function() {
            const garment = this.dataset.garment;
            const quantity = parseInt(this.value);
            updateQuantity(garment, quantity);
        });
        
        // Actualizar estado inicial
        const garment = input.dataset.garment;
        const quantity = parseInt(input.value);
        if (quantity > 0) {
            updateQuantity(garment, quantity);
        }
    });
});
</script>
{% endblock %}
''')

ORDER_DETAIL_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Detalles de la Orden #{{ order.order_number }}</h2>
    <div>
        <a href="/orders/{{ order.id }}/receipt" class="btn btn-info"> Generar Recibo</a>
        <a href="/" class="btn btn-secondary">← Volver</a>
    </div>
</div>

<div class="row">
    <div class="col-md-6">
        <div class="card">
            <div class="card-header">
                <h5 class="mb-0">Información de la Orden</h5>
            </div>
            <div class="card-body">
                <p><strong>Número de orden:</strong> {{ order.order_number }}</p>
                <p><strong>Cliente:</strong> {{ order.client_name or 'No especificado' }}</p>
                <p><strong>Teléfono:</strong> {{ order.phone or 'No especificado' }}</p>
                <p><strong>Estado:</strong> 
                    <span class="badge bg-{{ 'success' if order.status == 'listo' else 'warning' if order.status == 'pendiente' else 'secondary' }}">
                        {{ order.status }}
                    </span>
                </p>
                <p><strong>Fecha de creación:</strong> {{ order.created_at[:19] }}</p>
                <p><strong>Fecha de entrega:</strong> {{ order.delivery_date or 'No especificada' }}</p>
                <p><strong>Notas:</strong> {{ order.notes or 'Ninguna' }}</p>
            </div>
        </div>
        
        <div class="card mt-3">
            <div class="card-header">
                <h5 class="mb-0">Cambiar Estado</h5>
            </div>
            <div class="card-body">
                <form method="post" action="/orders/{{ order.id }}/status">
                    <div class="input-group">
                        <select class="form-select" name="status">
                            <option value="pendiente" {{ 'selected' if order.status == 'pendiente' }}>Pendiente</option>
                            <option value="proceso" {{ 'selected' if order.status == 'proceso' }}>En proceso</option>
                            <option value="listo" {{ 'selected' if order.status == 'listo' }}>Listo</option>
                            <option value="entregado" {{ 'selected' if order.status == 'entregado' }}>Entregado</option>
                        </select>
                        <button type="submit" class="btn btn-primary">Actualizar</button>
                    </div>
                </form>
            </div>
        </div>
    </div>
    
    <div class="col-md-6">
        <div class="card">
            <div class="card-header">
                <h5 class="mb-0"> Items de la Orden</h5>
            </div>
            <div class="card-body">
                <div class="table-responsive">
                    <table class="table table-striped">
                        <thead>
                            <tr>
                                <th>Prenda</th>
                                <th>Cantidad</th>
                                <th>Precio Unit.</th>
                                <th>Subtotal</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for item in items %}
                            <tr>
                                <td>{{ item.garment_type.title() }}</td>
                                <td>{{ item.quantity }}</td>
                                <td>${{ "%.2f"|format(item.unit_price) }}</td>
                                <td>${{ "%.2f"|format(item.subtotal) }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                        <tfoot>
                            <tr class="table-success">
                                <td colspan="3" class="text-end"><strong>TOTAL:</strong></td>
                                <td><strong>${{ "%.2f"|format(order.total or 0) }}</strong></td>
                            </tr>
                        </tfoot>
                    </table>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
''')

REPORTS_TEMPLATE = BASE_HTML.replace('{% block content %}{% endblock %}', '''
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h2> Reportes y Estadísticas</h2>
    <div>
        <a href="/export/orders.csv" class="btn btn-outline-success"> Exportar CSV</a>
        <a href="/export/orders.xlsx" class="btn btn-outline-primary"> Exportar Excel</a>
        {% if session.user_role == 'admin' %}
        <a href="/export/backup_all.zip" class="btn btn-outline-warning"> Backup Completo</a>
        {% endif %}
    </div>
</div>

<div class="row">
    <div class="col-md-6">
        <div class="card">
            <div class="card-header bg-success text-white">
                <h5 class="mb-0"> Ventas del Día</h5>
            </div>
            <div class="card-body text-center">
                <h3>${{ "%.2f"|format(sales_today) }}</h3>
                <p class="text-muted">Total vendido hoy</p>
            </div>
        </div>
    </div>
    
    <div class="col-md-6">
        <div class="card">
            <div class="card-header bg-info text-white">
                <h5 class="mb-0"> Servicios Más Populares</h5>
            </div>
            <div class="card-body">
                {% if popular %}
                <div class="list-group">
                    {% for service in popular %}
                    <div class="list-group-item d-flex justify-content-between align-items-center">
                        {{ service.garment_type.title() }}
                        <span class="badge bg-primary rounded-pill">{{ service.q }}</span>
                    </div>
                    {% endfor %}
                </div>
                {% else %}
                <p class="text-muted">No hay datos suficientes</p>
                {% endif %}
            </div>
        </div>
    </div>
</div>

<div class="card mt-4">
    <div class="card-header">
        <h5 class="mb-0"> Herramientas de Exportación</h5>
    </div>
    <div class="card-body">
        <div class="row">
            <div class="col-md-4 text-center">
                <h6> Exportar CSV</h6>
                <p class="text-muted">Formato compatible con Excel</p>
                <a href="/export/orders.csv" class="btn btn-outline-success">Descargar CSV</a>
            </div>
            <div class="col-md-4 text-center">
                <h6> Exportar Excel</h6>
                <p class="text-muted">Formato .xlsx avanzado</p>
                <a href="/export/orders.xlsx" class="btn btn-outline-primary">Descargar Excel</a>
            </div>
            {% if session.user_role == 'admin' %}
            <div class="col-md-4 text-center">
                <h6> Backup Completo</h6>
                <p class="text-muted">Respaldo de toda la base</p>
                <a href="/export/backup_all.zip" class="btn btn-outline-warning">Descargar Backup</a>
            </div>
            {% endif %}
        </div>
    </div>
</div>
{% endblock %}
''')

# ---------------------- EJECUCIÓN ----------------------
if __name__ == '__main__':
    print(" Sistema de Lavandería Effiwash iniciando...")
    print(" Accede en: http://localhost:5000")
    print(" Usuario: admin / Contraseña: admin123")
    app.run(debug=True, host='0.0.0.0', port=5000)